import requests
import datetime
import openpyxl

from time import sleep
from random import randint

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from solve_captcha.anticaptcha import solve_recaptchaV2

from clients.adszilla import AdsZilla
from working_data import anti_captcha_key, working_file, spin_link, miner_link
from general_functions.get_tot import get_totp


class HuobiZilla(AdsZilla):

    def __init__(self, mail: str, huobi_pass: str, ads_acc_num: int,
                 ads_acc_id: str, tot_secret=None, trc_address=None, access_key=None, secret_key=None):
        super().__init__(ads_acc_num, ads_acc_id)
        self.mail = mail
        self.huobi_pass = huobi_pass
        self.tot_secret = tot_secret
        self.trc_address = trc_address
        self.access_key = access_key
        self.secret_key = secret_key

    def __str__(self):
        return f"Email: {self.mail}, huobi pass: {self.huobi_pass}, tot: {self.tot_secret}, " \
               f"trc address: {self.trc_address}, access: {self.access_key}, secret: {self.secret_key}"

    def _check_if_logged_in(self):
        style = self.driver.find_element(By.XPATH,
                                         '/html/body/div[1]/div/section/div/header/div[2]/div[2]/div[2]').get_attribute(
            'style')
        if style != 'display: none;':
            return True
        else:
            return False

    def _solve_slider_captcha(self):
        WebDriverWait(self.driver, 120).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[4]/div/div/div[2]/div/div[2]/div[1]/div/div[1]/span')))
        slider = self.driver.find_element(By.XPATH,
                                          '/html/body/div[4]/div/div/div[2]/div/div[2]/div[1]/div/div[1]/span')
        ac = ActionChains(self.driver)
        ac.drag_and_drop_by_offset(slider, 360, 0).perform()

    def _send_login_and_password_press_login_button(self):
        account_field = self.driver.find_element(By.ID, 'login_name')
        account_field.clear()
        account_field.send_keys(self.mail)

        psw_field = self.driver.find_element(By.ID, 'password')
        psw_field.clear()
        psw_field.send_keys(self.huobi_pass)

        sleep(randint(2, 3))
        login_button_2 = self.driver.find_element(By.CLASS_NAME, 'login-btn-submit')
        login_button_2.click()

    def _check_if_captcha_appears(self):
        try:
            self.driver.find_element(By.CLASS_NAME, 'captcha-container')
            return True
        except NoSuchElementException:
            return False

    def _check_captcha_type(self):
        if self.driver.find_element(By.CLASS_NAME, 'gt-captcha').get_attribute('style') != "display: none;":
            return 'gt-captcha'
        if self.driver.find_element(By.CLASS_NAME, 'g-recaptcha').get_attribute('style') != "display: none;":
            return 'g-captcha'
        if self.driver.find_element(By.ID, 'alicaptcha-1').get_attribute('style') != "display: none;":
            return 'slider-captcha'

    def _switch_captcha(self):
        self.driver.find_element(By.CLASS_NAME, 'switch-btn').click()

    def _solve_huobi_captcha(self, captcha_type):
        if captcha_type == 'gt-captcha':
            self._switch_captcha()
            sleep(randint(1, 2))
            captcha_type = self._check_captcha_type()
            self._solve_huobi_captcha(captcha_type)
        elif captcha_type == 'slider-captcha':
            self._switch_captcha()
            sleep(randint(1, 2))
            captcha_type = self._check_captcha_type()
            self._solve_huobi_captcha(captcha_type)
        elif captcha_type == 'g-captcha':
            url = self.driver.current_url
            web_site_key = self.driver.find_element(By.CLASS_NAME, 'g-recaptcha').get_attribute('data-sitekey')
            g_resp = solve_recaptchaV2(key=anti_captcha_key, url=url, web_site_key=web_site_key)
            try:
                # in case g-captcha response field or callback function name changes, should change here also
                self.driver.execute_script("""document.getElementById("g-recaptcha-response").innerHTML=arguments[0];""",
                                           g_resp)
                self.driver.execute_script("""window.recaptchaSuccessCallback(arguments[0]);""", g_resp)
                sleep(randint(2, 3))
            except:
                print('ERROR: Ads account', self.acc_num, '- could not solve recaptcha')
        else:
            print('ERROR: Ads account', self.acc_num, '- unknown type of captcha')

    def _send_two_fa_key(self):
        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div/section/section/section/div[2]/form/div/div[2]/div/input')))
        totp = get_totp(self.tot_secret)
        two_fa_field = self.driver.find_element(
            By.XPATH, '/html/body/div[1]/div/section/section/section/div[2]/form/div/div[2]/div/input')
        two_fa_field.send_keys(totp)
        sleep(randint(1, 2))

    def login(self):
        self.driver.get("https://www.huobi.com/en-us/")
        self.close_all_pages_except_main()
        sleep(randint(1, 2))
        if self._check_if_logged_in():
            return

        login_button = self.driver.find_element(By.XPATH,
                                                "/html/body/div[1]/div/section/div/header/div[2]/div[2]/div[1]/a[1]")
        login_button.click()

        # wait for login button to appier
        try:
            WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
                (By.XPATH, '/html/body/div[1]/div/section/section/div/div[2]/form/div[3]/div[2]/a[2]')))
        except TimeoutException:
            print('ERROR: Ads account', self.acc_num, '- login button does not appear')

        try:
            self._send_login_and_password_press_login_button()
        except:
            print('ERROR: Ads account', self.acc_num, '- could not send login/pass data')

        sleep(randint(4, 5))

        if self._check_if_captcha_appears():
            captcha_type = self._check_captcha_type()
            self._solve_huobi_captcha(captcha_type)

        sleep(randint(1, 2))

        try:
            login_button_2 = self.driver.find_element(By.CSS_SELECTOR,'#__layout > section > section > div > div.login-inner > form > button.login-btn-submit')
            login_button_2.click()
        except:
            pass

        try:
            self._send_two_fa_key()
        except:
            print('ERROR: Ads account', self.acc_num, '- could not send 2fa key')

        return 'Logged in'

    def _bind2fa(self, wb, row, tot_col):
        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div/section/div/div[2]/div[3]/div/div[2]/div[2]/div[1]')))
        key = self.driver.find_element(By.XPATH,
                                       '/html/body/div[1]/div/section/div/div[2]/div[3]/div/div[2]/div[2]/div[1]')
        key_code = key.get_attribute('innerHTML').strip()
        print(self.mail, key_code)
        sheet = wb.active
        sheet.cell(row, column=tot_col).value = key_code
        wb.save(f"{working_file} back_up{datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}.xlsx")
        wb.close()
        return key_code

    def get_trading_amount(self):
        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div[1]/div/section/div[2]/div[1]/div/div[2]/div[2]/div/div[2]/div/div[2]/ul/li[1]/p/em')))
        trading_amount_filed = self.driver.find_element(By.XPATH,
                                                        '/html/body/div[1]/div/section/div[2]/div[1]/div/div[2]/div[2]/div/div[2]/div/div[2]/ul/li[1]/p/em').text
        trading_amount = trading_amount_filed.split(" ")[0]
        return int(trading_amount)

    def disable_email_verif(self):
        self.driver.get("https://account.huobi.com/en-us/user-center/security")
        self.close_all_pages_except_main()
        sleep(randint(2, 3))

        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div/section/div/div[2]/div[2]/div[1]/ul[2]/li[3]/div/button/span')))
        disable_email_button = self.driver.find_element(By.XPATH,
                                                        '/html/body/div[1]/div/section/div/div[2]/div[2]/div[1]/ul[2]/li[3]/div/button/span')
        disable_email_button.click()
        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div[1]/div/section/div/div[2]/div[2]/div[5]/div[1]/div/div[2]/form/div[1]/div/div[1]/span/span/span/button')))
        send_code = self.driver.find_element(By.XPATH,
                                             '/html/body/div[1]/div/section/div/div[2]/div[2]/div[5]/div[1]/div/div[2]/form/div[1]/div/div[1]/span/span/span/button')
        send_code.click()
        totp = get_totp(self.tot_secret)
        totp_field = self.driver.find_element(By.XPATH,
                                              '/html/body/div[1]/div/section/div/div[2]/div[2]/div[5]/div[1]/div/div[2]/form/div[2]/div/div/input')
        totp_field.send_keys(totp)

    def change_huobi_pass(self, new_psw):
        self.driver.get("https://account.huobi.com/en-us/user-center/security")
        self.close_all_pages_except_main()
        sleep(randint(2, 3))

        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div/section/div/div[2]/div[3]/div[1]/ul[1]/li[3]/button')))
        change_pass_button = self.driver.find_element(By.XPATH,
                                                      '/html/body/div[1]/div/section/div/div[2]/div[3]/div[1]/ul[1]/li[3]/button')
        change_pass_button.click()

        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div[1]/div/section/div/div[2]/div[3]/div[2]/div[1]/div/div[2]/form/div[1]/div/div/input')))
        old_pass = self.driver.find_element(By.XPATH,
                                            '/html/body/div[1]/div/section/div/div[2]/div[3]/div[2]/div[1]/div/div[2]/form/div[1]/div/div/input')
        new_pass = self.driver.find_element(By.XPATH,
                                            '/html/body/div[1]/div/section/div/div[2]/div[3]/div[2]/div[1]/div/div[2]/form/div[2]/div/div[1]/input')
        new_pass_2 = self.driver.find_element(By.XPATH,
                                              '/html/body/div[1]/div/section/div/div[2]/div[3]/div[2]/div[1]/div/div[2]/form/div[3]/div/div/input')

        old_pass.send_keys(self.huobi_pass)
        new_pass.send_keys(new_psw)
        new_pass_2.send_keys(new_psw)
        confirm_button = self.driver.find_element(By.XPATH,
                                                  '/html/body/div[1]/div/section/div/div[2]/div[3]/div[2]/div[1]/div/div[3]/div/button[2]')
        confirm_button.click()

    def set_2fa(self, wb, row, tot_col):
        self.driver.get('https://www.huobi.com/en-us/user_center/uc_auth/')
        self.close_all_pages_except_main()
        sleep(randint(2, 3))

        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div/section/section/section[2]/div/div/div[3]/div/div[1]/div[1]/dl/dd[2]/i')))
        verification = self.driver.find_element(By.XPATH,
                                                '/html/body/div[1]/div/section/section/section[2]/div/div/div[3]/div/div[1]/div[1]/dl/dd[2]/i')
        check = verification.get_attribute('class')
        if check == "hb_icon_toast_success":
            print("verified")
            self.driver.get('https://account.huobi.com/en-us/user-center/bind-ga')
            two_fa = self._bind2fa(wb, row, tot_col)
            code = get_totp(two_fa)

            print(code)

            WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
                (By.XPATH,
                 '/html/body/div[1]/div/section/div/div[2]/div[4]/div/form/div[1]/div/div/input')))

            code_field = self.driver.find_element(By.XPATH,
                                                  '/html/body/div[1]/div/section/div/div[2]/div[4]/div/form/div[1]/div/div/input')
            code_field.send_keys(code)

            submit_button = self.driver.find_element(By.XPATH,
                                                     '/html/body/div[1]/div/section/div/div[2]/div[4]/div/form/div[2]/div/button')
            submit_button.click()

            WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
                (By.XPATH,
                 '/html/body/div[1]/div/section/div/div[3]/div[1]/div/div[2]/form/div[1]/div/div[1]/span/span/span/button')))
            send_code = self.driver.find_element(By.XPATH,
                                                 '/html/body/div[1]/div/section/div/div[3]/div[1]/div/div[2]/form/div[1]/div/div[1]/span/span/span/button')
            send_code.click()

        else:
            print('account is not verified!!!!')
        return

    def get_trc_addr(self):
        self.driver.get("https://www.huobi.com/en-us/finance/deposit/usdt")
        self.close_all_pages_except_main()
        sleep(randint(2, 3))

        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/div[2]/div/div/div[2]')))
        trc_radio = self.driver.find_element(By.XPATH,
                                             '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/div[2]/div/div/div[2]')
        trc_radio.click()

        WebDriverWait(self.driver, 30).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/div[2]/section/div/section/div/button/span')))

        send_addr = self.driver.find_element(By.XPATH,
                                             '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/div[2]/section/div/section/div/button/span')
        send_addr.click()

        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[3]/div/div/div[2]/div/div[2]/div[2]')))
        confirm_but = self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div[2]/div/div[2]/div[2]')
        confirm_but.click()

        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/div[2]/section/div/section/div/div[2]/div[2]')))

        addr = self.driver.find_element(By.XPATH,
                                        '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/div[2]/section/div/section/div/div[2]/div[2]')
        trc_addr = addr.get_attribute('innerHTML').strip()
        print(self.mail, trc_addr)
        # sheet.cell(row, column=trc_col).value = trc_addr
        # wb.save(f"new200{datetime.now().strftime('%H%M%S')}.xlsx")
        # wb.close()

        # print(e, key_code)
        # sheet.cell(r, column=tot_col).value = key_code
        # wb.save(f"new200{datetime.now().strftime('%H%M%S')}.xlsx")

    def get_API_key(self):
        self.driver.get('https://www.huobi.com/en-us/apikey/')
        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div[1]/div/section/section/div[2]/div[1]/div/div[1]/section/div[2]/div/div[2]/label/input')))
        check_box1 = self.driver.find_element(By.XPATH,
                                              '/html/body/div[1]/div/section/section/div[2]/div[1]/div/div[1]/section/div[2]/div/div[2]/label/input')
        check_box2 = self.driver.find_element(By.XPATH,
                                              '/html/body/div[1]/div/section/section/div[2]/div[1]/div/div[1]/section/div[2]/div/div[3]/label/input')
        check_box1.click()
        check_box2.click()
        notes_field = self.driver.find_element(By.XPATH,
                                               '/html/body/div[1]/div/section/section/div[2]/div[1]/div/div[1]/section/div[1]/input')
        notes_field.send_keys('main2')
        create_button = self.driver.find_element(By.XPATH,
                                                 '/html/body/div[1]/div/section/section/div[2]/div[1]/div/div[1]/div/button/span[1]')
        create_button.click()

        try:
            WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
                (By.XPATH, '/html/body/div[4]/div/div/div[2]/section/div[1]/h3')))
            check_box3 = self.driver.find_element(By.XPATH,
                                                  '/html/body/div[4]/div/div/div[2]/section/div[2]/div[1]/div/div[1]/label/input')
            check_box4 = self.driver.find_element(By.XPATH,
                                                  '/html/body/div[4]/div/div/div[2]/section/div[2]/div[1]/div/div[2]/label/input')
            check_box3.click()
            check_box4.click()
            i_understand_button = self.driver.find_element(By.XPATH,
                                                           '/html/body/div[4]/div/div/div[2]/section/div[3]/div/button[2]')
            i_understand_button.click()
        except:
            # there is no consent field
            pass

        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[4]/div/div/div[2]/section/div[2]/form/dl/dd/div/div[1]/input')))
        two_fa_filed = self.driver.find_element(By.CSS_SELECTOR, '#ga')
        two_fa_filed.send_keys(get_totp(self.tot_secret))

        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[4]/div/div/div[2]/section/div[3]/div/div/button/span[1]')))
        confirm_button1 = self.driver.find_element(By.XPATH,
                                                   '/html/body/div[4]/div/div/div[2]/section/div[3]/div/div/button/span[1]')
        confirm_button1.click()

        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[4]/div/div/div[2]/section/div[2]/div/div[2]/div/div')))
        self.access_key = self.driver.find_element(By.XPATH,
                                                   '/html/body/div[4]/div/div/div[2]/section/div[2]/div/div[2]/div/div').text
        self.secret_key = self.driver.find_element(By.XPATH,
                                                   '/html/body/div[4]/div/div/div[2]/section/div[2]/div/div[3]/div/div').text
        confirm_button2 = self.driver.find_element(By.XPATH,
                                                   '/html/body/div[4]/div/div/div[2]/section/div[3]/div/div/button/span[1]')
        confirm_button2.click()

        print(self.access_key, '/////', self.secret_key)

        return self.access_key, self.secret_key

    def save_api_data_to_file(self, row, column_a, column_s, access_key=None, secret_key=None):
        wb = openpyxl.load_workbook(working_file)
        sheet = wb.active
        sheet.cell(row + 1, column=column_a).value = self.access_key if self.access_key else access_key
        sheet.cell(row + 1, column=column_s).value = self.secret_key if self.secret_key else secret_key
        wb.save(working_file)
        wb.close()

    def check_if_registered(self):
        try:
            self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div/div[2]/section/div[3]/div/button')
            return True
        except:
            return False

    def use_miner(self):
        self.redirect_to_page(miner_link)
        sleep(randint(4, 5))
        self._check_if_have_to_spin()
        mining_amount = self._get_available_mining_amount()
        while mining_amount > 0:
            start_mining_button = self.driver.find_element(By.XPATH,
                                                           '/html/body/div[1]/div/section/div[2]/div/div[2]/div[2]/button[2]/span')
            start_mining_button.click()
            WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
                (By.XPATH, '/html/body/div[4]/div/div/div[2]/div')))
            sleep(randint(5, 7))
            result = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div/div[2]/div/div[1]').text
            print(self.acc_num, ' got ', result)
            close_pop_up_button = self.driver.find_element(By.XPATH, '/html/body/div[4]/div/div/div[2]/div/button[1]')
            close_pop_up_button.click()
            sleep(randint(2, 3))

            mining_amount = self._get_available_mining_amount()

    def _get_available_mining_amount(self):
        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div/section/div[2]/div/div[2]/div[1]/div[1]/em')))
        mining_amount_text = self.driver.find_element(By.XPATH,
                                                      '/html/body/div[1]/div/section/div[2]/div/div[2]/div[1]/div[1]/em').text
        mining_amount = int(mining_amount_text.split(' ')[0])
        return mining_amount

    def _check_if_have_to_spin(self):
        try:
            self.driver.switch_to.frame(self.driver.find_element(By.NAME, 'mining'))
            spin_button = self.driver.find_element(By.CLASS_NAME, 'go-btn')
            spin_button.click()
            WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
                (By.XPATH, '/html/body/div[1]/main/div/div[1]/div/div[6]')))
            start_mining_button = self.driver.find_element(By.XPATH, '/html/body/div[1]/main/div/div[1]/div/div[6]')
            start_mining_button.click()
            sleep(randint(2, 3))
        except:
            return

    def turn_on_point_cards(self):
        sleep(randint(2, 3))
        try:
            WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
                (By.XPATH, "/html/body/div[1]/div/section/div[2]/div[2]/div[2]/div/div[1]/div[2]/span")))

            turn_on_button = self.driver.find_elements(By.CLASS_NAME, 'switch-wrap')[0].find_element(By.CLASS_NAME, 'switch-ui')

            self.driver.execute_script("arguments[0].setAttribute('class','switch-ui switch-ui-middle switch-ui-on')", turn_on_button)
            WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
                (By.XPATH, "/html/body/div[4]/div/div/div[2]/div/div[3]/div/button/span[1]")))

            confirm_button = self.driver.find_element(By.XPATH, "/html/body/div[4]/div/div/div[2]/div/div[3]/div/button/span[1]")
            confirm_button.click()

        except:
            return

    def use_spin(self):
        self.redirect_to_page(spin_link)
        try:
            WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
                            (By.XPATH, '/html/body/div[1]/div/section/div[2]/div[3]/div[1]/div[1]/dl/dd[9]/button')))
            sleep(2)
            spin_button = self.driver.find_element(By.XPATH, '/html/body/div[1]/div/section/div[2]/div[3]/div[1]/div[1]/dl/dd[9]/button')
            spin_button.click()
        except:
            pass

    def withdraw_USDT(self, adrs, network):
        self.redirect_to_page('https://www.huobi.com/en-us/finance/withdraw/usdt/')
        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/section/dl[3]/dd/div/div[2]/div[2]/label/span')))

        input_data = False

        try:
            if network == 'trc':
                trc_radio = self.driver.find_element(By.XPATH,
                                               '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/section/dl[3]/dd/div/div[2]/div[2]/label/span')
                trc_radio.click()

            if network == 'sol':
                sol_radio = self.driver.find_element(By.XPATH,
                                               '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/section/dl[3]/dd/div/div[6]/div[2]/label/span')
                sol_radio.click()

            address = self.driver.find_element(By.XPATH, '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/section/section[1]/div[2]/div[1]/div/textarea')
            address.send_keys(adrs)

            sleep(1)
            all_button = self.driver.find_element(By.XPATH, '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/section/dl[4]/dd/div/div[2]/div/span/button')
            all_button.click()

            input_data = True

        except:
            pass

        try:
            sleep(7)
            check_box1 = self.driver.find_element(By.XPATH,
                                             '/html/body/div[3]/div/div/div[2]/section/div[2]/div/div/div/ul/li[1]/div/div/ul/li/label/i')
            check_box2 = self.driver.find_element(By.XPATH,
                                              '/html/body/div[3]/div/div/div[2]/section/div[2]/div/div/div/ul/li[2]/div/div/ul/li/label/i')
            check_box3 = self.driver.find_element(By.XPATH,
                                              '/html/body/div[3]/div/div/div[2]/section/div[2]/div/div/div/ul/li[3]/div/div/ul/li/label/i')
            check_box4 = self.driver.find_element(By.XPATH,
                                                  '/html/body/div[3]/div/div/div[2]/section/div[2]/div/div/p[2]/div/div/ul/li/label/i')
            check_box1.click()
            check_box2.click()
            check_box3.click()
            check_box4.click()

            confirm = self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div[2]/section/div[3]/div/button/span')
            confirm.click()

        except:
            pass

        if not input_data:
            if network == 'trc':
                trc_radio = self.driver.find_element(By.XPATH,
                                                     '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/section/dl[3]/dd/div/div[2]/div[2]/label/span')
                trc_radio.click()

            if network == 'sol':
                sol_radio = self.driver.find_element(By.XPATH,
                                               '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/section/dl[3]/dd/div/div[7]/div[2]/label/span')
                sol_radio.click()

            if network == 'algo':
                algo_radio = self.driver.find_element(By.XPATH,
                                               '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/section/dl[3]/dd/div/div[6]/div[2]/label/span')
                algo_radio.click()

            sleep(1)
            all_button = self.driver.find_element(By.XPATH,
                                                  '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/section/dl[4]/dd/div/div[2]/div/span/button')
            all_button.click()

        withdraw_button = self.driver.find_element(By.XPATH, '/html/body/div[1]/div/section/main/article/div/section[1]/div[1]/div[2]/section/section[2]/button/span')
        withdraw_button.click()

        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[3]/div/div/div[2]/section/div[1]/a')))

        consent_check = self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div[2]/section/div[3]/div/label/input')
        consent_check.click()

        confirm = self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div[2]/section/div[4]/button[2]/span')
        confirm.click()

        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[3]/div/div/div[2]/section/div[2]/form/dl/dd/div/div[1]/input')))

        two_fa_field = self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div[2]/section/div[2]/form/dl/dd/div/div[1]/input')
        two_fa_field.send_keys(get_totp(self.tot_secret))

        confirm_two = self.driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div[2]/section/div[3]/div/div/button/span[1]')
        confirm_two.click()
