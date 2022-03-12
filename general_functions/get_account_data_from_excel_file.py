import openpyxl

from working_data import data_columns, working_file


def get_account_data_from_file(row):
    wb = openpyxl.load_workbook(working_file)
    sheet = wb.active

    row += 1

    mail = sheet.cell(row, column=data_columns['mail']).value
    huobi_pass = sheet.cell(row, column=data_columns['huobi_pass']).value
    if mail is None or huobi_pass is None:
        return
    mail_pass = sheet.cell(row, column=data_columns['mail_pass']).value
    ads_acc_num = sheet.cell(row, column=data_columns['ads_acc_num']).value
    ads_acc_id = sheet.cell(row, column=data_columns['ads_acc_id']).value
    tot = sheet.cell(row, column=data_columns['tot']).value
    ads_group = sheet.cell(row, column=data_columns['ads_group']).value
    trc_addr = sheet.cell(row, column=data_columns['trc']).value
    access_key = sheet.cell(row, column=data_columns['access_key']).value
    secret_key = sheet.cell(row, column=data_columns['secret_key']).value
    proxy_ip = sheet.cell(row, column=data_columns['proxy_ip']).value
    proxy_user = sheet.cell(row, column=data_columns['proxy_user']).value

    try:
        proxy = {'http': 'http://' + proxy_user[1:] + "@" + proxy_ip,
                 'https': 'http://' + proxy_user[1:] + "@" + proxy_ip}

    except:
        # there is no proxy data
        proxy = {}

    huobi_acc_data = {
        'mail': mail,
        'mail_pass': mail_pass,
        'huobi_pass': huobi_pass,
        'ads_acc_num': ads_acc_num,
        'ads_acc_id': ads_acc_id,
        'tot': tot,
        'ads_group': ads_group,
        'trc': trc_addr,
        'access_key': access_key,
        'secret_key': secret_key,
        'proxy': proxy,
    }

    wb.close()

    return huobi_acc_data
